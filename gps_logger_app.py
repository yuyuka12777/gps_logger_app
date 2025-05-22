import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import serial
import pynmea2
import threading
import datetime
import os
from openpyxl import Workbook, load_workbook
import serial.tools.list_ports
import csv
import keyboard

class GPSLoggerApp:
    def __init__(self, master):
        self.master = master
        self.master.title("位置情報取得君")
        self.running = False
        self.ser = None
        self.history = []
        self.save_format = self.ask_save_format()
        self.filename = self.get_filename()
        self.save_key_var = tk.StringVar(value="F2")
        self.save_key_entry = None
        self.global_hotkey = None
        self.setup_widgets()
        self.check_existing_file()
        self.master.protocol("WM_DELETE_WINDOW", self.on_closing)

    def ask_save_format(self):
        # 保存形式を最初に選択
        answer = messagebox.askquestion("保存形式の選択", "CSVで保存しますか？（いいえの場合はExcelで保存）")
        return "csv" if answer == "yes" else "excel"

    def get_filename(self):
        today = datetime.datetime.now().strftime("%Y_%m_%d")
        ext = "csv" if getattr(self, "save_format", "excel") == "csv" else "xlsx"
        return f"{today}.{ext}"

    def check_existing_file(self):
        if os.path.exists(self.filename):
            res = messagebox.askyesno("ファイル確認", f"{self.filename} が既に存在します。続けますか？")
            if not res:
                i = 1
                base, ext = os.path.splitext(self.filename)
                while True:
                    new_name = f"{base}_{i}{ext}"
                    if not os.path.exists(new_name):
                        self.filename = new_name
                        break
                    i += 1

    def setup_widgets(self):
        frame = ttk.Frame(self.master)
        frame.pack(padx=10, pady=10)

        # 利用可能なCOMポート一覧を取得
        ports = [port.device for port in serial.tools.list_ports.comports()]
        if not ports:
            ports = ["COM1"]
        ttk.Label(frame, text="ポート名:").grid(row=0, column=0)
        self.port_var = tk.StringVar(value=ports[0])
        self.port_combo = ttk.Combobox(frame, textvariable=self.port_var, values=ports, state="readonly")
        self.port_combo.grid(row=0, column=1)
        ttk.Label(frame, text="ボーレート:").grid(row=1, column=0)
        self.baud_entry = ttk.Entry(frame)
        self.baud_entry.insert(0, "4800")
        self.baud_entry.grid(row=1, column=1)

        # 履歴保存キー設定
        ttk.Label(frame, text="履歴保存キー:").grid(row=2, column=0)
        self.save_key_entry = ttk.Entry(frame, textvariable=self.save_key_var)
        self.save_key_entry.grid(row=2, column=1)
        self.save_key_entry.bind("<KeyPress>", self.on_save_key_press)
        self.start_btn = ttk.Button(frame, text="開始", command=self.start)
        self.start_btn.grid(row=3, column=0, pady=5)
        self.stop_btn = ttk.Button(frame, text="停止", command=self.stop, state=tk.DISABLED)
        self.stop_btn.grid(row=3, column=1, pady=5)
        self.save_btn = ttk.Button(frame, text="履歴をコピー", command=self.copy_history)
        self.save_btn.grid(row=4, column=0, columnspan=2, pady=5)
        self.history_box = tk.Text(frame, width=50, height=15)
        self.history_box.grid(row=5, column=0, columnspan=2, pady=5)
        author_label = ttk.Label(frame, text="by yuyuka12777", foreground="gray")
        author_label.grid(row=6, column=0, columnspan=2, pady=(10, 0))

        self.bind_save_key()

    def on_save_key_press(self, event):
        self.save_key_var.set(event.keysym)
        self.bind_save_key() 
        return "break"

    def bind_save_key(self):
        try:
            self.master.unbind_all('<Key>')
            keystr = self.save_key_var.get().strip()
            if not keystr:
                return
            event_key = f"<{keystr}>"
            
            try:
                self.master.bind(event_key, lambda e: self.save_current_position())
            except tk.TclError as e:
                print(f"キーバインドエラー: {e}")
        
            # 特殊キー名をkeyboardライブラリ用に変換
            kb_keystr = keystr.lower()
            kb_key_mapping = {
                'shift_l': 'shift',
                'shift_r': 'shift',
                'control_l': 'ctrl',
                'control_r': 'ctrl', 
                'alt_l': 'alt',
                'alt_r': 'alt',
                'return': 'enter',
                'escape': 'esc'
            }

            if kb_keystr in kb_key_mapping:
                kb_keystr = kb_key_mapping[kb_keystr]
                
            # グローバルホットキー設定
            try:
                if self.global_hotkey:
                    keyboard.remove_hotkey(self.global_hotkey)
                if len(kb_keystr) == 1 or kb_keystr.startswith('f') or kb_keystr in ['space', 'enter', 'esc', 'tab']:
                    self.global_hotkey = keyboard.add_hotkey(kb_keystr, self.save_current_position)
                    print(f"グローバルホットキー登録: {kb_keystr}")
                else:
                    print(f"サポート外のグローバルホットキー: {kb_keystr}")
                    self.global_hotkey = None
            except Exception as e:
                print(f"グローバルホットキー登録エラー: {e}")
                self.global_hotkey = None
        except Exception as e:
            print(f"bind_save_keyエラー: {e}")

    def start(self):
        port = self.port_var.get()
        baud = int(self.baud_entry.get())
        try:
            self.ser = serial.Serial(port, baud, timeout=1)
            self.running = True
            self.start_btn.config(state=tk.DISABLED)
            self.stop_btn.config(state=tk.NORMAL)
            threading.Thread(target=self.read_gps, daemon=True).start()
        except Exception as e:
            messagebox.showerror("エラー", f"シリアルポート接続失敗: {e}")

    def stop(self):
        self.running = False
        if self.ser:
            self.ser.close()
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)

    def read_gps(self):
        while self.running:
            try:
                line = self.ser.readline().decode('ascii', errors='replace')
                if line.startswith('$GNGGA'):
                    msg = pynmea2.parse(line)
                    self.current_lat = msg.latitude
                    self.current_lon = msg.longitude
                    self.current_time = datetime.datetime.now().strftime("%H:%M:%S")
            except Exception:
                continue

    def save_current_position(self):
        if hasattr(self, 'current_lat') and hasattr(self, 'current_lon'):
            entry = f"{self.current_time} 緯度: {self.current_lat}, 経度: {self.current_lon}"
            self.history.append([self.current_time, self.current_lat, self.current_lon])
            self.history_box.insert(tk.END, entry + "\n")
            self.write_to_file(self.current_time, self.current_lat, self.current_lon)
        else:
            messagebox.showinfo("情報", "まだGPSデータを取得していません。")

    def copy_history(self):
        self.master.clipboard_clear()
        self.master.clipboard_append(self.history_box.get("1.0", tk.END))

    def write_to_file(self, time, lat, lon):
        if self.save_format == "csv":
            file_exists = os.path.exists(self.filename)
            with open(self.filename, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                if not file_exists:
                    writer.writerow(["時刻", "緯度", "経度"])
                writer.writerow([time, lat, lon])
        else:
            if os.path.exists(self.filename):
                wb = load_workbook(self.filename)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(["時刻", "緯度", "経度"])
            ws.append([time, lat, lon])
            wb.save(self.filename)

    def on_closing(self):
        if self.global_hotkey:
            try:
                keyboard.remove_hotkey(self.global_hotkey)
            except Exception:
                pass
        if self.ser:
            self.ser.close()
        self.master.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = GPSLoggerApp(root)
    root.mainloop()